import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import io, base64

# ==========================================================
# 🎨 Tampilan Aplikasi
# ==========================================================
st.set_page_config(page_title="Aplikasi Konversi Format Anjab ke SIASN", layout="centered")

st.markdown("""
<style>
.stApp {
    background: linear-gradient(135deg, #0057b7 0%, #ffd700 100%);
    background-attachment: fixed;
}
.title {
    text-align: center;
    color: #003366;
    font-size: 30px;
    font-weight: 800;
    margin-bottom: 5px;
    text-shadow: 1px 1px 2px rgba(255,255,255,0.8);
}
.subtitle {
    text-align: center;
    color: #1e293b;
    font-size: 16px;
    margin-bottom: 25px;
    font-weight: 700;
    letter-spacing: 1px;
}
.card {
    background-color: rgba(255,255,255,0.95);
    padding: 25px 30px;
    border-radius: 15px;
    border: 2px solid rgba(0,87,183,0.3);
    box-shadow: 0 3px 10px rgba(0,0,0,0.1);
    margin-bottom: 25px;
}
.footer {
    text-align: center;
    margin-top: 40px;
    font-size: 13px;
    color: #002244;
    font-weight: 600;
}
.flag-id {
    display:inline-block;width:28px;height:18px;
    background:linear-gradient(to bottom,#ff0000 50%,#ffffff 50%);
    border:1px solid #666;border-radius:2px;margin:0 2px;
    vertical-align:middle;
}
.flag-ua {
    display:inline-block;width:28px;height:18px;
    background:linear-gradient(to bottom,#0057b7 50%,#ffd700 50%);
    border:1px solid #666;border-radius:2px;margin:0 2px;
    vertical-align:middle;
}
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="title">🗂️ APLIKASI KONVERSI FORMAT ANJAB KE SIASN</div>', unsafe_allow_html=True)
st.markdown('<div class="subtitle">#SLAVAUKRAINI</div>', unsafe_allow_html=True)

# ==========================================================
# ⚙️ UTILITAS
# ==========================================================
def load_template_workbook_from_file(path_b64="template_base64.txt"):
    with open(path_b64, "r", encoding="utf-8-sig") as f:
        b64data = f.read()
    data = base64.b64decode(b64data)
    bio = io.BytesIO(data)
    return load_workbook(bio)

def get_first_sheet(df_dict):
    first_sheet_name = list(df_dict.keys())[0]
    return df_dict[first_sheet_name], first_sheet_name

# ==========================================================
# 🔍 FUNGSI EKSTRAKSI CERDAS
# ==========================================================
def extract_single_value(df, label):
    rows, cols = df.shape
    label_lower = label.lower().strip()
    for r in range(rows):
        for c in range(cols):
            val = df.iat[r, c]
            if pd.isna(val): continue
            if label_lower in str(val).lower():
                for offset in range(1, 6):
                    if c + offset < cols and str(df.iat[r, c + offset]).strip() == ":":
                        if c + offset + 1 < cols:
                            return str(df.iat[r, c + offset + 1]).strip()
                for offset in range(1, 6):
                    if c + offset < cols and pd.notna(df.iat[r, c + offset]):
                        v = str(df.iat[r, c + offset]).strip()
                        if v and v != ":": return v
    return None

def extract_multi_value_smart_last(df, label, right_offset, down_offset, log_msgs=None):
    rows, cols = df.shape
    found=[]
    for r in range(rows):
        for c in range(cols):
            v=df.iat[r,c]
            if pd.isna(v):continue
            s=str(v).strip().lower()
            if s==label.lower() or s.startswith(label.lower()):
                found.append((r,c))
    if not found:
        if log_msgs: log_msgs.append(f"⚠️ Label '{label}' tidak ditemukan.")
        return []
    r,c=found[-1]
    start_row, start_col=r+down_offset, c+right_offset
    results, first_found=[],False
    for rr in range(start_row,rows):
        if start_col>=cols:break
        v=df.iat[rr,start_col]
        if not first_found:
            if pd.isna(v) or str(v).strip()=="": continue
            first_found=True
        if first_found:
            if pd.isna(v) or str(v).strip()=="": break
            results.append(str(v).strip())
    if log_msgs: log_msgs.append(f"✅ '{label}' ditemukan (baris {r+1}, kolom {c+1}) {len(results)} data.")
    return results

def extract_tugas_pokok_multi_smart(df,log_msgs=None):
    label="tugas pokok"
    rows,cols=df.shape
    for r in range(rows):
        for c in range(cols):
            v=df.iat[r,c]
            if pd.isna(v):continue
            if str(v).strip().lower()==label:
                start_row=r+3
                offsets={"A":c+3,"C":c+6,"D":c+7,"E":c+9,"F":c+8}
                results={k:[] for k in offsets.keys()}
                for key,col_idx in offsets.items():
                    first_found=False
                    for rr in range(start_row,rows):
                        val=df.iat[rr,col_idx]
                        if not first_found:
                            if pd.isna(val) or str(val).strip()=="": continue
                            first_found=True
                        if first_found:
                            if pd.isna(val) or str(val).strip()=="": break
                            results[key].append(str(val).strip())
                if log_msgs: log_msgs.append(f"✅ 'Tugas Pokok' ditemukan (baris {r+1}, kolom {c+1}).")
                return results["A"],results["C"],results["D"],results["E"],results["F"]
    return [],[],[],[],[]

def extract_bahan_kerja(df,log_msgs=None):
    rows,cols=df.shape
    label_row,label_col=None,None
    for r in range(rows):
        for c in range(cols):
            v=df.iat[r,c]
            if pd.isna(v):continue
            s=str(v).strip().lower()
            if s=="8":
                right_val=df.iat[r,c+1] if c+1<cols else ""
                if str(right_val).strip().lower()=="bahan kerja":
                    label_row,label_col=r,c+1
    if label_row is None:
        if log_msgs: log_msgs.append("⚠️ Label '8 Bahan Kerja' tidak ditemukan.")
        return []
    start_row = label_row + 2
    start_col = label_col + 3   # kolom F
    results, first_found = [], False
    for rr in range(start_row, rows):
        if start_col >= cols: break
        v = df.iat[rr, start_col]
        if not first_found:
            if pd.isna(v) or str(v).strip() == "":
                continue
            first_found = True
        if first_found:
            if pd.isna(v) or str(v).strip() == "":
                break
            results.append(str(v).strip())
    if log_msgs:
        log_msgs.append(f"✅ 'Bahan Kerja' ditemukan (baris {label_row+1}, kolom F) dan {len(results)} data diambil.")
    return results

def extract_perangkat_kerja(df, log_msgs=None):
    rows, cols = df.shape
    label_row, label_col = None, None
    for r in range(rows):
        for c in range(cols):
            v = df.iat[r, c]
            if pd.isna(v): continue
            s = str(v).strip().lower()
            if s == "9":
                next_val = df.iat[r, c + 1] if c + 1 < cols else ""
                if str(next_val).strip().lower() == "perangkat kerja":
                    label_row, label_col = r, c + 1
    if label_row is None:
        if log_msgs: log_msgs.append("⚠️ Label '9 Perangkat Kerja' tidak ditemukan.")
        return []
    start_row = label_row + 2
    start_col = label_col + 3   # Kolom F
    results, first_found = [], False
    for rr in range(start_row, rows):
        if start_col >= cols: break
        v = df.iat[rr, start_col]
        if not first_found:
            if pd.isna(v) or str(v).strip() == "":
                continue
            first_found = True
        if first_found:
            if pd.isna(v) or str(v).strip() == "":
                break
            results.append(str(v).strip())
    if log_msgs:
        log_msgs.append(f"✅ 'Perangkat Kerja' ditemukan (baris {label_row+1}, kolom F) dan {len(results)} data diambil.")
    return results

# ==========================================================
# 🧩 ISIAN DEFAULT INFOJAB I
# ==========================================================
def apply_defaults(ws,jabatan):
    jabatan_options={
        "Penelaah Teknis Kebijakan":("S-1/Sarjana","S-1 SEMUA JURUSAN",7),
        "Pengolah Data dan Informasi":("Diploma III/Sarjana Muda","D-III SEMUA JURUSAN",6),
        "Pengadministrasi Perkantoran":("SLTA","SLTA SEDERAJAT",5),
    }
    c4,d4,z4=jabatan_options[jabatan]
    ws["C4"],ws["D4"],ws["E4"]=c4,d4,"ADMINISTRASI PERKANTORAN"
    ws["F4"]="minimal sesuai syarat jabatan pada Standar Kompetensi Jabatan"
    ws["G4"]="Mampu menyusun, mengonsep, menganalisis, melaporkan dan evaluasi suatu bidang data berkaitan dengan tugas jabatan"
    defaults={
        "H4":"Bakat Verbal","H5":"Intelegensia","H6":"Bakat Numerik","H7":"Bakat Ketelitian",
        "H8":"Koordinasi Motorik","H9":"Kecekatan Jari","H10":"Koordinasi Mata, Tangan, Kaki",
        "I4":"Konvensional","I5":"Realistik","I6":"Investigasi","I7":"Sosial",
        "J4":"Directing Control Planning (DCP)","J5":"Feeling-Idea-Fact (FIF)",
        "J6":"Influencing (INFLU)","J7":"Sensory & Judgmental Creteria (SJC)",
        "J8":"Measurable and Verifiable Creteria (MVC)","J9":"Dealing with People (DEPL)",
        "J10":"Repetitive and Continuous (REPCON)","J11":"Performing Under Stress (PUS)",
        "J12":"Set of Limits, Tolerance and Other Standart (STS)","J13":"Variety and Changing Conditions (VARCH)",
        "K4":"Berdiri","K5":"Berjalan","K6":"Duduk","K7":"Bekerja dengan jari","K8":"Berbicara",
        "K9":"Mendengar","K10":"Melihat","L4":"Laki-laki/Perempuan","M4":"Tegap","N4":"58",
        "O4":"ramah dan sopan","P4":"tidak ada syarat khusus","Q4":"tidak ada syarat khusus berat badan",
        "R4":"N","T4":"Memadukan data","T5":"Mengkoordinasi data","T6":"Menganalisis data",
        "T7":"Menyusun data","T8":"Menghitung data","T9":"Menyalin data","T10":"Membandingkan data",
        "U4":"Tidak ada","V4":"Menerima instruksi","W4":"Baik","Z4":z4
    }
    for cell,val in defaults.items():
        ws[cell]=val

# ==========================================================
# 🖥️ ANTARMUKA STREAMLIT
# ==========================================================
st.markdown('<div class="card">', unsafe_allow_html=True)
jabatan=st.selectbox("🧩 Pilih Nama Jabatan",[
    "Penelaah Teknis Kebijakan",
    "Pengolah Data dan Informasi",
    "Pengadministrasi Perkantoran",
])
file=st.file_uploader("📁 Unggah FILE SUMBER.xlsx",type=["xlsx"])
st.markdown("</div>", unsafe_allow_html=True)

if file:
    st.info(f"📄 File sumber diterima: {file.name}")
    df_dict=pd.read_excel(file,sheet_name=None,header=None)
    df,sheet_name=get_first_sheet(df_dict)
    st.info(f"✅ Menggunakan hanya sheet pertama: **{sheet_name}**")

    if st.button("🚀 Proses & Buat File Output"):
        with st.spinner("🔄 Memproses data..."):
            wb=load_template_workbook_from_file()
            ws1,ws2=wb["INFOJAB I"],wb["INFOJAB II"]
            log=[]
            val=extract_single_value(df,"IKTISAR JABATAN")
            if val: ws1["B4"]=val; log.append("✅ IKTISAR JABATAN diisi.")
            else: log.append("⚠️ IKTISAR JABATAN tidak ditemukan.")
            apply_defaults(ws1,jabatan)
            tj=extract_multi_value_smart_last(df,"Tanggung Jawab",3,2,log)
            ww=extract_multi_value_smart_last(df,"Wewenang",3,2,log)
            for i,v in enumerate(tj,4): ws1[f"X{i}"]=v
            for i,v in enumerate(ww,4): ws1[f"Y{i}"]=v
            tp_a,tp_c,tp_d,tp_e,tp_f=extract_tugas_pokok_multi_smart(df,log)
            for i,v in enumerate(tp_a,4): ws2[f"A{i}"]=v
            for i,v in enumerate(tp_c,4): ws2[f"C{i}"]=v
            for i,v in enumerate(tp_d,4): ws2[f"D{i}"]=v
            for i,v in enumerate(tp_e,4): ws2[f"E{i}"]=v
            for i,v in enumerate(tp_f,4): ws2[f"F{i}"]=v
            hk=extract_multi_value_smart_last(df,"Hasil Kerja",3,2,log)
            for i,v in enumerate(hk,4): ws2[f"B{i}"]=v
            bk=extract_bahan_kerja(df,log)
            for i,v in enumerate(bk,4): ws2[f"G{i}"]=v
            pk=extract_perangkat_kerja(df,log)
            for i,v in enumerate(pk,4): ws2[f"H{i}"]=v
            buf=io.BytesIO(); wb.save(buf); buf.seek(0)
            out_name=file.name.rsplit(".",1)[0]+"-CONVERTED.xlsx"
            st.download_button("💾 Unduh Hasil Konversi",buf,out_name,mime="application/vnd.ms-excel")
        for msg in log:
            if "✅" in msg: st.success(msg)
            elif "⚠️" in msg: st.warning(msg)
            else: st.info(msg)

# ==========================================================
# 🖥️ FOOTER 🇮🇩 × 🇺🇦
# ==========================================================
st.markdown('''
<div class="footer">
<span class="flag-id"></span> × <span class="flag-ua"></span>
&nbsp;Dibuat dengan cinta ❤️ oleh&nbsp;
<b>Tim Anjab Bagian Organisasi Kabupaten Muaro Jambi</b>
</div>
''', unsafe_allow_html=True)
